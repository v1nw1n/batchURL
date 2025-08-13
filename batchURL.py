import os
import io
import time
import base64
import argparse
from bs4 import BeautifulSoup, Comment
import threading
import logging
from queue import Queue
from datetime import datetime
from urllib.parse import urlsplit, urlunsplit ,urlparse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage, ImageOps
import socket
import dns.resolver
import requests
from requests.exceptions import RequestException, ProxyError, ConnectTimeout
from selenium.webdriver.support.ui import WebDriverWait
#===日志配置===
os.makedirs(os.path.join(os.getcwd(), 'log'), exist_ok=True)
LOG_FILENAME = "./log/batchURL.log"
with open(LOG_FILENAME, "w", encoding="utf-8") as f:
    f.write('')

logging.basicConfig(
    filename=LOG_FILENAME,
    filemode="a",
    format="%(asctime)s [%(levelname)s] %(message)s",
    level=logging.INFO,
    encoding="utf-8"
)

class FirefoxRequestFilter(logging.Filter):
    def filter(self, record):
        msg = record.getMessage()
        FILTER_WORD = [".firefox.", ".mozilla.", ".google.",".ico",".js", ".css", ".png", ".jpg", ".jpeg", ".gif", ".svg", ".ico", ".woff", ".woff2", ".ttf", ".eot", ".otf"]
        if any(domain in msg.lower() for domain in FILTER_WORD):
            return False
        return True

# seleniumwire logger
logger = logging.getLogger("seleniumwire")
logger.setLevel(logging.INFO)
logger.addFilter(FirefoxRequestFilter())

# 给根logger的handler也添加过滤器，确保写文件的handler经过过滤
root_logger = logging.getLogger()
for handler in root_logger.handlers:
    handler.addFilter(FirefoxRequestFilter())
#========
from AISupport import *
from selenium.webdriver.firefox.service import Service
from seleniumwire import webdriver  
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.common.exceptions import WebDriverException, TimeoutException
from domainE import extract_main_domain
from WXWorkWebHook import taskFinishNotify
from NetIOHelper import NetIOHelper
# === 配置 ===
THREAD_MAX_LIMIT = 29
TARGET_IMG_HIGHT = 200
ROW_HEIGHT = 150
PROJ_INDEX = datetime.now().strftime("%Y%m%d%H%M%S")
normal_count = 0
abnormal_count = 0
unreachable_count = 0

def print_banner():
    banner = "Cl9fX19fXyAgICAgICBfICAgICAgIF8gICAgIF8gICBfX19fX19fIF8gICAgIAp8IF9fXyBcICAgICB8IHwgICAgIHwgfCAgIHwgfCB8IHwgX19fIHwgfCAgICAKfCB8Xy8gLyBfXyBffCB8XyBfX198IHxfXyB8IHwgfCB8IHxfLyB8IHwgICAgCnwgX19fIFwvIF9gIHwgX18vIF9ffCAnXyBcfCB8IHwgfCAgICAvfCB8ICAgIAp8IHxfLyB8IChffCB8IHx8IChfX3wgfCB8IHwgfF98IHwgfFwgXHwgfF9fX18KXF9fX18vIFxfXyxffFxfX1xfX198X3wgfF98XF9fXy9cX3wgXF9cX19fX18vCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgLS1ieSB2MW53MW4gICAgICAgICAgICAgICAgCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIAo="
    print(base64.b64decode(banner.encode('utf-8')).decode('utf-8'))


# 页面类型判断
def extract_visible_text_and_count(html):
    soup = BeautifulSoup(html, 'html.parser')
    #记录标签总数（原始 HTML）
    tag_count = len(soup.find_all(True))
    #删除特定标签
    for tag in soup(['script', 'style', 'head', 'meta', 'noscript']):
        tag.decompose()
    #删除 HTML 注释
    for comment in soup.find_all(string=lambda text: isinstance(text, Comment)):
        comment.extract()
    #提取可见文本（清理后）
    visible_text = soup.get_text(separator=' ', strip=True)

    return str(soup).lower(), visible_text.lower(), tag_count


def page_judge(url: str, html: str,title: str ,http_status: int, imgPath: str, token: str = None) -> str:
    res = None
    if token :
        res =  page_judge_ai(imgPath,token)
    if res is None:
        return page_judge_local(url=url,html=html,title=title,http_status=http_status)
    else:
        return res


def page_judge_ai(imgPath,token):
    imgPath = imgTokenSimplizer(imgPath)
    res = agent_call(token=token,imgPath = imgPath,text= PROMPT_PAGE_JUDGE)
    logging.info(f"AI call->{res}:{imgPath}:prompt->PROMPT_PAGE_JUDGE")
    res = getAIResponse(res)
    if res in TYPE_DEFINE.keys() :
        return TYPE_DEFINE[res]["type"]
    else:
        logging.error(f"AI 响应异常:{res}")
        print(f"AI 异常,转本地判断,详情查看日志")
    return None


def page_judge_local(url: str, html: str,title: str ,http_status: int) -> str:
    html_lower, visible_text, tag_count = extract_visible_text_and_count(html)
    TAG_THRESHOLD = 30
    KEYWORDS = {
        "welcome": ["tomcat", "nginx", "openresty"],
        "error": [
            "400", "bad request",
            "401", "unauthorized", 
            "403", "forbidden",
            "404", "not found", 
            "405", "method not allowed",
            "500", "internal server error",
            "502", "bad gateway",
            "503", "service unavailable",
            
        ],
        "login": ["login","type=\"password\"","登录"],
        "ui_error": ["error"]
    }
    #错误页 http status 4xx 5xx
    if http_status is not None and http_status >= 400:
        logging.info(f"http status code -> {http_status} : {url}")
        return TYPE_DEFINE["3"]["type"]
    #登录页 命中关键词 and 结构复杂
    for keyword in KEYWORDS["login"]:
        if keyword in html_lower:
            logging.info(f"命中关键词 -> {keyword} : 标签数 -> {tag_count} : {url} ")
            return TYPE_DEFINE["2"]["type"]
    if tag_count >= TAG_THRESHOLD:
        #异常系统 命中关键词 and 结构复杂
        for keyword in KEYWORDS["ui_error"]:
            if keyword in html_lower:
                logging.info(f"命中关键词 -> {keyword} : 标签数 -> {tag_count} : {url} ")
                return TYPE_DEFINE["6"]["type"]
    else:
        #白页 结构简单 and 可见文本为空
        if not visible_text.strip() :
            logging.info(f"可见文本为空 ; 标签数 -> {tag_count} : {url}")
            return TYPE_DEFINE["5"]["type"]
        #错误页 命中关键词 and 结构简单
        for keyword in KEYWORDS["error"]:
            if keyword in html_lower:
                logging.info(f"命中关键词 -> {keyword} : 标签数 -> {tag_count} : {url} ")
                return TYPE_DEFINE["3"]["type"]
    #欢迎页 命中关键词 and 结构简单 and title包含关键词welcom
    for keyword in KEYWORDS["welcome"]:
        if keyword in html_lower and "welcom" in title.lower():
            logging.info(f"命中关键词 -> {keyword} : title -> {title} : 标签数 -> {tag_count} : {url} ")
            return TYPE_DEFINE["4"]["type"]
    
    #默认正常系统
    return TYPE_DEFINE["1"]["type"]


SCREENSHOTS_DIR = ".\\screenshots\\"

def resize_image(image_bytes, target_height=TARGET_IMG_HIGHT,idx = None):
    img = PILImage.open(io.BytesIO(image_bytes))

    if idx is not None and isinstance(idx,int):
        os.makedirs(SCREENSHOTS_DIR, exist_ok=True)
        img.save(os.path.join(SCREENSHOTS_DIR, f"{PROJ_INDEX}_{idx}.png"), format="PNG", optimize=True)
    h_percent = target_height / float(img.size[1])
    w_size = int(img.size[0] * h_percent)
    img = img.resize((w_size, target_height), PILImage.LANCZOS)
    img = ImageOps.expand(img, border=2, fill='black')
    buffer = io.BytesIO()
    img.save(buffer, format="PNG", optimize=True)
    buffer.seek(0)
    return buffer


image_save_queue = Queue()

def image_saver_worker():
    while True:
        item = image_save_queue.get()
        if item is None:
            break
        idx, image_bytes = item
        try:
            img = PILImage.open(io.BytesIO(image_bytes))
            os.makedirs(SCREENSHOTS_DIR, exist_ok=True)
            img.save(os.path.join(SCREENSHOTS_DIR, f"{PROJ_INDEX}_{idx}.png"), format="PNG", optimize=True)
        except Exception as e:
            logging.warning(f"截图保存失败 idx={idx}: {e}")
        image_save_queue.task_done()


def cleanup_screenshots():
    screenshot_dir = os.path.abspath(SCREENSHOTS_DIR)
    deleted_count = 0
    for file in os.listdir(screenshot_dir):
        if file.startswith(f"{PROJ_INDEX}_") and file.endswith(".png"):
            try:
                os.remove(os.path.join(screenshot_dir, file))
                deleted_count += 1
            except Exception as e:
                logging.warning(f"无法删除截图 {file}: {e}")
    print(f"\n🧹  已删除截图文件数量：{deleted_count}")


def net_check(target_url, proxy_address, timeout=5):
    result = {
        "target": target_url,
        "proxy_status": None,
        "direct_status": None,
        "proxy_code": None,
        "direct_code": None,
        "result": None
    }
    proxies = {
        "http": proxy_address,
        "https": proxy_address
    }
    # 直连访问
    try:
        direct_resp = requests.get(target_url, timeout=timeout)
        result["direct_status"] = True
        result["direct_code"] = direct_resp.status_code
    except RequestException:
        result["direct_status"] = False

    # 代理访问
    try:
        proxy_resp = requests.get(target_url, proxies=proxies, timeout=timeout)
        result["proxy_status"] = True
        result["proxy_code"] = proxy_resp.status_code
    except (ProxyError, ConnectTimeout, RequestException):
        result["proxy_status"] = False
    # 逻辑判断
    ps = result["proxy_status"]
    ds = result["direct_status"]
    pc = result["proxy_code"]
    dc = result["direct_code"]

    if ps and ds:
        if pc == dc:
            result["result"] = "任意网络"
        elif pc in [403, 405, 502] and dc < 400:
            result["result"] = "直连访问"
        else:
            result["result"] = "直连访问【需确认】"
    elif ps and not ds:
        result["result"] = "代理访问"
    elif not ps and ds:
        result["result"] = "直连访问"
    else:
        result["result"] = "不可访问"
    logging.info(f"proxy_status->{result['proxy_status']}:direct_status->{result['direct_status']}:proxy_code->{result['proxy_code']}:direct_code->{result['direct_code']}:{target_url}")
    return result["result"]

def resolve_ip(url, custom_dns=None):
    domain = urlparse(url).hostname
    try:
        if custom_dns:
            resolver = dns.resolver.Resolver()
            resolver.nameservers = [custom_dns]
            answer = resolver.resolve(domain)
            logger.info(f"dns->{custom_dns}:ip->{ip}:{url}")
            return answer[0].to_text()
        else:
            ip = socket.gethostbyname(domain)
            logger.info(f"dns->default:ip->{ip}:{url}")
            return ip
    except:
        if custom_dns:
            try:
                ip = socket.gethostbyname(domain)
                logger.info(f"dns->default:ip->{ip}:{url}")
                return ip
            except:
                logger.info(f"dns->default:ip->无法解析:{url}")
                return ""
        logger.info(f"dns->default:ip->无法解析:{url}")
        return ""


# 获取状态码 
def normalize_url(url: str):
    parts = urlsplit(url)
    # 丢弃 fragment
    return urlunsplit((parts.scheme, parts.netloc, parts.path, parts.query, ''))


def get_status_code(driver):
    try:
        norm_current = normalize_url(driver.current_url)
        for request in reversed(driver.requests):
            if request.response :
                norm_request = normalize_url(request.url)
                if  norm_request  == norm_current:
                    logging.info(f"获取 HTTP 响应码: {request.response.status_code}:{ request.url}")
                    return request.response.status_code
    except Exception as e:
        logging.exception(f"获取 HTTP 响应码失败: {str(e)}:{ request.url}")
        pass
    return -1


def get_firefox_exe_dir():
    return "C:\\Program Files\\Mozilla Firefox\\firefox.exe"

    for path in os.environ.get('PATH', '').split(os.pathsep):
        exe_path = os.path.join(path, 'firefox.exe')
        if os.path.isfile(exe_path):
            return exe_path
    common_paths = [
        r"C:\Program Files\Mozilla Firefox\firefox.exe",
        r"C:\Program Files (x86)\Mozilla Firefox\firefox.exe",
        os.path.expandvars(r"%LOCALAPPDATA%\Mozilla Firefox\firefox.exe"),
        os.path.expandvars(r"%PROGRAMFILES%\Mozilla Firefox\firefox.exe"),
        os.path.expandvars(r"%PROGRAMFILES(x86)%\Mozilla Firefox\firefox.exe"),
    ]
    for p in common_paths:
        if os.path.isfile(p):
            return p
    return "C:\\Program Files\\Mozilla Firefox\\firefox.exe"



# 创建浏览器实例 
def create_browser():
    options = FirefoxOptions()
    options.add_argument("--headless")
    #禁用HTTP2
    options.set_preference("network.http.spdy.enabled", False)
    options.set_preference("network.http.spdy.enabled.http2", False)
    options.set_preference("network.http.spdy.enabled.deps", False)
    options.set_preference("network.http.spdy.websockets", False)
    options.set_preference("network.http.http2.enabled", False)
    #======
    options.accept_insecure_certs = True
    options.binary_location = get_firefox_exe_dir()
    service = Service(os.environ.get('geckodriver_exe'))
    driver = webdriver.Firefox( seleniumwire_options={'disable_capture': False}, options=options,service=service )
    driver.set_page_load_timeout(15)
    return driver


#  浏览器池工作线程 
def worker(thread_id, task_queue, result_dict, lock, llm_token, progress_callback, status_dict ,ns_dns, proxy_url):
    driver = create_browser()
    net_helper = NetIOHelper(proxy_url=proxy_url, custom_dns=ns_dns)
    while True:
        try:
            task = task_queue.get(timeout=3)
        except:
            break

        idx, url = task
        status_dict["current"] = f"线程-{thread_id} 正在处理: {idx} - {url}"

        title = ""
        html = ""
        image = None
        topDoamin =  extract_main_domain(url)

        try:
            ip = net_helper.resolve_ip(url)
            proxy_reachable = net_helper.net_check(url)
            # proxy_reachable = net_check(url,proxy_url)
            # ip = resolve_ip(url,ns_dns)
        except Exception as e:
            logging.exception(f"处理 {url} 异常:{e} ")

        try:
            driver.get(url)
            WebDriverWait(driver, 6).until(
                lambda d: d.execute_script('return document.readyState') == 'complete'
            )
            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            if soup.title and soup.title.string:
                title = soup.title.string.strip()
            http_status = get_status_code(driver)
            screenshot = driver.get_screenshot_as_png()
            image_save_queue.put((idx, screenshot))
            image = resize_image(image_bytes = screenshot,idx = idx)
            imgPath = os.path.join( os.path.abspath(SCREENSHOTS_DIR), f"{PROJ_INDEX}_{idx}.png")
            logging.info(f"imgPath->{imgPath}:{url}")
            status = page_judge(url=url, html = html,title = title, http_status = http_status, imgPath = imgPath, token = llm_token)
            
        except TimeoutException:
            status = "无法访问(访问超时)"
            image = None
        except WebDriverException:
            status = "无法访问(Web异常)"
            image = None
        except Exception as e:
            status = "无法访问(其他异常)"
            image = None
            logging.exception(f"处理 {url} 异常:{e} ")

        with lock:
            result_dict[idx] = {
                "id": idx,
                "url": url,
                "status": status,
                "page_title": title,
                "IP": ip,
                "topDomain": topDoamin,
                "proxy_reachable": proxy_reachable,
                "image": image
            }
            if progress_callback:
                progress_callback()

        task_queue.task_done()
    net_helper.close()
    driver.quit()



# 写入 Excel 文件
def write_excel(results: list, output_path):
    global normal_count, abnormal_count, unreachable_count 
    wb = Workbook()
    ws_ok = wb.active
    ws_ok.title = "优选目标"
    ws_error = wb.create_sheet(title="错误页")
    ws_failed = wb.create_sheet(title="无法访问")

    headers = ["ID", "URL", "访问状态", "标题", "域名解析IP", "主域名", "网络访问", "截图"]

    for ws in [ws_ok, ws_error, ws_failed]:
        ws.append(headers)
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 10
        ws.column_dimensions["H"].width = 50
    
    for res in results:
        row_data = [
            res["id"],
            res["url"],
            res["status"],
            res.get("page_title", ""),
            res.get("IP", ""),
            res.get("topDomain", ""),
            res.get("proxy_reachable", "")
        ]
        image = res.get("image")
        status = res.get("status", "")

        # 判断目标写入哪个 sheet
        if status.startswith("无法访问"):
            ws = ws_failed
            unreachable_count += 1
        elif "错误" in status:
            ws = ws_error
            abnormal_count += 1
        else:
            ws = ws_ok
            normal_count += 1

        row_num = ws.max_row + 1
        for col_num, value in enumerate(row_data, start=1):
            ws.cell(row=row_num, column=col_num, value=value)
        if image:
            img = XLImage(image)
            ws.add_image(img, f"H{row_num}")
            ws.row_dimensions[row_num].height = ROW_HEIGHT
        else:
            ws.cell(row=row_num, column=8, value="")
            ws.row_dimensions[row_num].height = 20
    
    wb.save(output_path)
    print(f"\n✅  Excel 文件已保存: {output_path}")
    return output_path


# 自动计算线程数
def calculate_worker_count(url_count, max_limit=THREAD_MAX_LIMIT):
    if url_count <= 10:
        return min(2, url_count)
    elif url_count <= 250:
        return min(10, url_count)
    elif url_count <= 400:
        return min(20, url_count)
    else:
        return min(max_limit, url_count // 20 + 2)

output_excel = ""
def main():
    print_banner()
    parser = argparse.ArgumentParser(description="从URLs获取更多访问信息")
    parser.add_argument('-i', '--input', metavar = "file",default='url-prod.txt', help='定义目标(需要http(s)前缀),一行一个目标(txt)')
    parser.add_argument('-o', '--output', metavar = "file name",default=f'batchURL_{PROJ_INDEX}.xlsx', help='定义输出文件名，不加后缀')
    parser.add_argument('-t','--llm-token',metavar="token", help='开启AI支持,配置token')
    parser.add_argument('-u','--friend-ui', action='store_true', help='是否启用进度条展示(默认关闭)')
    parser.add_argument('-c','--clean-img', action='store_true', help='程序结束后删除所有本地截图(默认保留)')
    parser.add_argument('-p','--net-check',metavar = "proxy",help='配置一个无害的代理,检查目标的网络可达性,仅支持http(s)')
    parser.add_argument('-n','--ns',metavar = "DNS address" ,help='指定用于解析URL对应域名的 DNS 服务器,默认使用系统NS')
    parser.add_argument('-r','--push-res',action='store_true' ,help='是否推送任务完成通知到企业微信机器人(默认关闭)')

    args = parser.parse_args()


    input_file = args.input
    output_excel = f"{args.output}_{PROJ_INDEX}.xlsx"
    llm_token = args.llm_token
    use_progress_bar = args.friend_ui
    is_clean_img = args.clean_img


    gecko_path = os.environ.get('geckodriver_exe')
    if not gecko_path or not os.path.exists(gecko_path):
        print(f"❌  环境变量 'geckodriver_exe' 未设置或路径无效: {gecko_path}")
        return
    
    if not os.path.exists(input_file):
        print(f"❌  未找到输入文件：{input_file}")
        return

    with open(input_file, 'r', encoding='utf-8') as f:
        urls = [line.strip().strip('\'"') for line in f if line.strip()]
    url_count = len(urls)
    if url_count == 0:
        print("❌  输入 URL 为空")
        return

    llm_token_check = False
    if llm_token is not None:
        if is_token_valid(llm_token):
            output_excel = f"{args.output}_{PROJ_INDEX}_AI.xlsx"
            llm_token_check = True
            print("✅  LLM TOKEN已配置")
        else :
            llm_token = None
            print("⚠️  LLM TOKEN不可用")
        
    worker_count = calculate_worker_count(url_count)
    print(f"📊  总计 URL: {url_count}，浏览器池线程数: {worker_count}")


    task_queue = Queue()
    for idx, url in enumerate(urls, start=1):
        task_queue.put((idx, url))

    results_dict = {}
    lock = threading.Lock()
    threads = []
    status_dict = {"current": ""}
    #进度信息
    progress_count = [0]
    progress_bar = None
    if use_progress_bar:
        try:
            from tqdm import tqdm
            progress_bar = tqdm(total=url_count, desc="⏸️  处理进度", ncols=80)
        except ImportError:
            print("⚠️  未安装 tqdm,进度条自动切换为轻量模式")
            use_progress_bar = False

    def progress_callback():
        progress_count[0] += 1
        if use_progress_bar and progress_bar:
            progress_bar.update(1)
        else:
            print(f"\r⏸️  进度: {progress_count[0]}/{url_count}", end="")

    start_time = time.time()
    for i in range(worker_count):
        t = threading.Thread(target=worker, args=(i + 1, task_queue, results_dict, lock, llm_token, progress_callback, status_dict, args.ns, args.net_check))
        t.start()
        threads.append(t)
    image_thread = threading.Thread(target=image_saver_worker)
    image_thread.start()
    for t in threads:
        t.join()
    image_save_queue.put(None)
    image_thread.join()

    if progress_bar:
        progress_bar.close()
    end_time = time.time()

    results = [results_dict[i] for i in sorted(results_dict.keys())]
    fileP = write_excel(results, output_excel)
    
    if not use_progress_bar:
        duration = end_time - start_time
        print(f"\n⏱️  处理完毕，总耗时：{duration:.2f} 秒（约 {duration / 60:.2f} 分钟）")
    if is_clean_img:
        cleanup_screenshots()
    else:
        print(f"\n📂  截图文件已保存在：{os.path.abspath(SCREENSHOTS_DIR)}")
    if llm_token_check:
        getTokenDeal()
    #任务完成推送消息
    if args.push_res:
        taskFinishNotify({"projectName": args.output,"normal_count":normal_count,"abnormal_count":abnormal_count,"unreachable_count":unreachable_count}, fileP, "1")

if __name__ == "__main__":
    main()

#TODO: 程序异常/崩溃/用户终止时：删除已产生的文件，清理遗留的进程（崩溃进程/浏览器进程）